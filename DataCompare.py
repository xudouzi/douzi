#encoding=utf-8
#影院信息的更新
import unittest
import sys
reload(sys)
sys.setdefaultencoding('utf-8')
from excel import *
import time
import csv
import codecs
import xlrd
import string
import xlsxwriter
import xlwt
from openpyxl.styles import Alignment

from openpyxl.styles import PatternFill
datapath_reference = r"/Users/douzi/Desktop/ge/副本5013724-MASSIVE_MIMO_FMC——焊接bom (1).xlsx"
sheetname_reference = r"MASSIVE_MIMO_FMC"

datapath_nonreference = r"/Users/douzi/Desktop/ge/MASSIVE_MIMO_FMC——焊接bom20190618.xlsx"
sheetname_nonreference ="MASSIVE_MIMO_FMC"
col = 2

def sheet_name(dataPath,sheetname):
    parseExcel = PsrseExcel()
    parseExcel.loasworkBook(dataPath)
    sheet = parseExcel.getSheetByName(sheetname)
    return sheet,parseExcel


#excel 表中读取位号
def read_weihao_excel(rowNo,colNo,datapath,sheetname):
    item=sheet_name(datapath,sheetname)
    sheet = item[0]
    parseExcel = item[1]
    data=parseExcel.getCellValue(sheet,rowNo=rowNo,colNo=colNo)
    return data

def max_row(datapath,sheetname):
    item = sheet_name(datapath,sheetname)
    sheet = item[0]
    parseExcel = item[1]
    max_row = parseExcel.getRowEndVumber(sheet)
    return max_row

def max_col(datapath, sheetname):
    item = sheet_name(datapath, sheetname)
    sheet = item[0]
    parseExcel = item[1]
    max_col = parseExcel.getColEndNuber(sheet)
    #print"max_clo", max_col
    return max_col


def write_different_data(dataPath,sheetname,row,col,content):
    item = sheet_name(dataPath,sheetname)
    sheet = item[0]
    parseExcel = item[1]
    data_item = parseExcel.writecell(sheet,content,row,col)
    print data_item


def writecell(self,sheet,content,rowNo = None,colNo = None):
    print "content",content
    print "rowNo",rowNo
    print "colNo",colNo

    sheet.cell(row=rowNo,column=colNo).value = content

def compareData():
    reference_row = max_row(datapath_reference,sheetname_reference)
    nonreference_row = max_row(datapath_nonreference,sheetname_nonreference)
    print reference_row,nonreference_row
    if reference_row !=nonreference_row:
        print "有数据的最大行数不相等，取小的行数，检查没有数据的最大行数是否填写了其他提示文字，或者线框有缺失，若有请删除 保持最大的行数一致"
    max_row_excel = nonreference_row
    max_col_excel = max_col(datapath_nonreference,sheetname_nonreference)

    for i in range(1, max_row_excel):
        data_non = []
        data_ref = []
        non_excel = read_weihao_excel(i+1, col, datapath_nonreference, sheetname_nonreference).replace('，', ',').split(',')
        ref_excel = read_weihao_excel(i+1, col, datapath_reference, sheetname_reference).replace('，', ',').split(',')
        for j in range(len(non_excel)):
            data_non.append(str(non_excel[j]))

        for k in range(len(ref_excel)):
            data_ref.append(str(ref_excel[k]))

        item = list(set(non_excel).difference(set(ref_excel)))
        data = openpyxl.load_workbook(datapath_nonreference)
        sheetnames = data.get_sheet_names()
        table = data.get_sheet_by_name(sheetnames[0])
        table = data.active
        if item !=[]:
            print i+1,item
            item_join= ','.join(item)
            table.cell(i+1,max_col_excel+1).value = item_join
            table.cell(i + 1, max_col_excel + 1).fill =PatternFill("solid", fgColor="FF0000")
            data.save(datapath_nonreference)



if __name__ == "__main__":
    compareData()










#获取有数据的最大的行数

