#encoding=utf-8

import openpyxl
#写excel
class PsrseExcel(object):
    def __init__(self):
        self.workbook = None
        self.excelFile = None

    def loasworkBook(self,excelpathName):
        try:
            self.workbook = openpyxl.load_workbook(excelpathName)
        except Exception ,e:
            raise e

    def getSheetByName(self,sheetName):
        try:
            sheet = self.workbook.get_sheet_by_name(sheetName)
            return sheet
        except Exception,e:
            raise e

    def getRowEndVumber(self,sheet):
        return sheet.max_row

    def getColEndNuber(self,sheet):
        return sheet.max_column

    def getRowStartNumber(self,sheet):
        return sheet.min_row

    def getColStartNumber(self,sheet):
        return sheet.min_column

    def getRow(self,sheet,rowNo):
        return list(sheet.rows)[rowNo-1]

    def getColumn(self,sheet,colNo):
        return list(sheet.column)[colNo-1]

    def getCellValue(self,sheet,coordinate= None,rowNo = None,colNo = None):
        if coordinate != None:
            try:
                return sheet.cell(coordinate=coordinate).value
            except Exception,e:
                raise e
        elif coordinate is None and rowNo is not None and colNo is not None:
            try:
                return sheet.cell(row=rowNo,column= colNo).value
            except Exception,e:
                raise e
        else:
            raise Exception,e

    def writecell(self,sheet,content,rowNo = None,colNo = None):
        print "content",content
        print "rowNo",rowNo
        print "colNo",colNo

        sheet.cell(row=rowNo,column=colNo).value = content




