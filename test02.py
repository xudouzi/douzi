#encoding=utf-8
import openpyxl
data = openpyxl.load_workbook('/Users/douzi/Desktop/excel_test.xlsx')
print(data.get_named_ranges()) # 输出工作页索引范围
print(data.get_sheet_names()) # 输出所有工作页的名称
# 取第一张表
sheetnames = data.get_sheet_names()
table = data.get_sheet_by_name(sheetnames[0])
table = data.active
print(table.title) # 输出表名
nrows = table.max_row # 获得行数
ncolumns = table.max_column # 获得行数
values =["我要你" ,"哈哈哈","事实上"]
for i in values:
    print i
    table.cell(1,1).value = i

data.save('/Users/douzi/Desktop/excel_test.xlsx')
